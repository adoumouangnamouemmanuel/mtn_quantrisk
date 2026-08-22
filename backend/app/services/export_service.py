"""Export service — generates PDF board reports and Excel KRI data files."""
import io
import logging
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..services.scenario_service import get_all_kpis
from ..services.brief_service import list_board_briefs

logger = logging.getLogger(__name__)

# ── Excel KRI Export ──────────────────────────────────────────────────────────


def export_kri_excel(period: str | None = None) -> bytes:
    """Export the KRI register to an Excel workbook.

    Returns the file content as bytes (caller writes to response).
    """
    kpis = get_all_kpis(period)

    wb = Workbook()
    ws = wb.active
    ws.title = "KRI Register"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    # Status colors
    status_fills = {
        "Safe": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        "Watch": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "Warning": PatternFill(start_color="FED7AA", end_color="FED7AA", fill_type="solid"),
        "Critical": PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid"),
    }

    # Headers
    headers = ["KPI ID", "Name", "Category", "Unit", "FY25 Value",
               "Lower Threshold", "Upper Threshold", "Status", "Trend (24m)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for row_idx, kpi in enumerate(kpis, 2):
        values = [
            kpi.get("id", ""),
            kpi.get("name", ""),
            kpi.get("category", ""),
            kpi.get("unit", ""),
            kpi.get("fy25Value"),
            kpi.get("lowerThreshold"),
            kpi.get("upperThreshold"),
            kpi.get("currentStatus", ""),
            " → ".join(str(v) for v in (kpi.get("trend24m") or [])[-6:]),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            # Color-code status column
            if col == 8 and val in status_fills:
                cell.fill = status_fills[val]

    # Auto-fit column widths
    for col in range(1, len(headers) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col).value or ""))
            for r in range(1, len(kpis) + 2)
        )
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 40)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Write to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── PDF Board Brief Export ────────────────────────────────────────────────────


def export_brief_pdf(brief_id: str | None = None) -> bytes:
    """Export board brief(s) to a PDF document.

    If brief_id is provided, exports that specific brief.
    Otherwise exports the most recent brief.
    """
    from fpdf import FPDF

    briefs = list_board_briefs()
    if not briefs:
        raise ValueError("No board briefs available. Generate one first.")

    if brief_id:
        brief = next((b for b in briefs if b.get("id") == brief_id), None)
        if not brief:
            raise ValueError(f"Brief {brief_id} not found")
    else:
        brief = briefs[0]  # Most recent

    class BriefPDF(FPDF):
        def header(self):
            self.set_font("Helvetica", "B", 10)
            self.set_text_color(100, 100, 100)
            self.cell(0, 8, "MTN Ghana — QuantRisk Board Brief", align="L")
            self.cell(0, 8, datetime.now(timezone.utc).strftime("%d %b %Y"), align="R", new_x="LMARGIN", new_y="NEXT")
            self.line(10, self.get_y(), 200, self.get_y())
            self.ln(4)

        def footer(self):
            self.set_y(-15)
            self.set_font("Helvetica", "I", 8)
            self.set_text_color(150, 150, 150)
            self.cell(0, 10, f"Page {self.page_no()}/{{nb}} — Confidential", align="C")

    pdf = BriefPDF()
    pdf.alias_nb_pages()
    pdf.add_page()

    # Title
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(20, 20, 20)
    pdf.cell(0, 12, brief.get("title", "Board Risk Brief"), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    # Metadata
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(100, 100, 100)
    meta_parts = [
        f"Scenarios: {', '.join(brief.get('scenarioIds', []))}",
        f"Generated: {brief.get('generatedAt', 'N/A')}",
        f"Severity Score: {brief.get('severityScore', 'N/A')}",
    ]
    est = brief.get("estimatedImpact", {})
    if est:
        meta_parts.append(f"Est. Impact: {est.get('magnitude', '')} {est.get('currency', '')}{est.get('unit', '')}")
    pdf.cell(0, 6, "  |  ".join(meta_parts), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(6)

    # Executive Summary
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_text_color(30, 30, 30)
    pdf.cell(0, 8, "Executive Summary", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(50, 50, 50)
    pdf.multi_cell(0, 5, brief.get("executiveSummary", "No summary available."))
    pdf.ln(4)

    # Key KPI Impacts
    kpi_impacts = brief.get("keyKpiImpacts", [])
    if kpi_impacts:
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_text_color(30, 30, 30)
        pdf.cell(0, 8, "Key KPI Impacts", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 10)
        for impact in kpi_impacts:
            kpi_id = impact.get("kpiId", "")
            narrative = impact.get("narrative", "")
            pdf.set_font("Helvetica", "B", 10)
            pdf.cell(0, 6, f"  {kpi_id}", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Helvetica", "", 9)
            pdf.set_text_color(70, 70, 70)
            pdf.multi_cell(0, 4.5, f"    {narrative}")
            pdf.ln(1)
            pdf.set_text_color(50, 50, 50)
        pdf.ln(2)

    # Recommended Actions
    actions = brief.get("recommendedActions", [])
    if actions:
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_text_color(30, 30, 30)
        pdf.cell(0, 8, "Recommended Actions", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 10)
        for i, action in enumerate(actions, 1):
            pdf.cell(0, 6, f"  {i}. {action}", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(4)

    # Calibration Notes
    cal_notes = brief.get("calibrationNotes", "")
    if cal_notes:
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_text_color(30, 30, 30)
        pdf.cell(0, 8, "Calibration Notes", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(50, 50, 50)
        pdf.multi_cell(0, 5, cal_notes)

    # Write to buffer
    buf = io.BytesIO()
    pdf_output = pdf.output()
    if isinstance(pdf_output, str):
        buf.write(pdf_output.encode("latin-1"))
    else:
        buf.write(pdf_output)
    buf.seek(0)
    return buf.read()


# ── Excel Scenario Comparison Export ──────────────────────────────────────────


def export_scenario_comparison_excel(
    scenario_a: dict, scenario_b: dict, results_a: dict, results_b: dict
) -> bytes:
    """Export a scenario comparison to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Scenario Comparison"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    headers = ["KPI", "Base Value", f"A: {scenario_a.get('name', '')}",
               f"B: {scenario_b.get('name', '')}", "Delta A", "Delta B", "Worse"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Merge results by kpiId
    results_a_map = {r["kpiId"]: r for r in results_a.get("results", [])}
    results_b_map = {r["kpiId"]: r for r in results_b.get("results", [])}
    all_kpi_ids = sorted(set(list(results_a_map.keys()) + list(results_b_map.keys())))

    for row_idx, kpi_id in enumerate(all_kpi_ids, 2):
        ra = results_a_map.get(kpi_id, {})
        rb = results_b_map.get(kpi_id, {})
        base = ra.get("baseValue") or rb.get("baseValue", "")
        val_a = ra.get("scenarioValue", "")
        val_b = rb.get("scenarioValue", "")
        delta_a = ra.get("deltaPct", "")
        delta_b = rb.get("deltaPct", "")
        worse = "A" if (delta_a or 0) < (delta_b or 0) else "B" if (delta_b or 0) < (delta_a or 0) else "Tie"

        for col, val in enumerate([kpi_id, base, val_a, val_b, delta_a, delta_b, worse], 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
